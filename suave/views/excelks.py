import re
import hashlib
import openpyxl
from django.db import transaction
from django.contrib.auth.models import User
from benefisiariu.models import Benefisiariu, AddressTL
from kni.models import Business, Program, LocBussiness, Employee, Finance
from suave.models import (
    CreditInfo, FinancialAssessment, ProductService, 
    MainCustomer, Competitor, MarketAssessment
)
from custom.models import (
    Municipality, AdministrativePost, Village, Status, 
    TIpu_Programa, Sector, Faze, Year
)

# ══════════════════════════════════════════════════════════════════
# MAPPING & HELPERS
# ══════════════════════════════════════════════════════════════════

MUNICIPALITY_ALIASES = {
    'liquiça': 'liquiçá', 'liquica': 'liquiçá', 'liquisa': 'liquiçá',
    'likisa': 'liquiçá', 'liqisa': 'liquiçá', 'dili': 'dili',
    'baucau': 'baucau', 'bobonaro': 'bobonaro', 'ainaro': 'ainaro',
    'ermera': 'ermera', 'covalima': 'covalima', 'lautem': 'lautem',
    'manatuto': 'manatuto', 'manufahi': 'manufahi', 'viqueque': 'viqueque',
    'aileu': 'aileu', 'atauro': 'dili', 'oe-cusse': 'regiao administrativa especial oe-cusse ambeno',
}

SECTOR_ALIASES = {
    'turizmu': 'turismu', 'turismo': 'turismu', 'agricultura': 'agrikultura',
    'agriculture': 'agrikultura', 'indusria': 'industria', 'industri': 'industria',
    'komersiu': 'komersiu', 'comercio': 'komersiu', 'teknolojia': 'teknolojia',
}

STATUS_MAP = {
    'Ativu': 'Ativu', 'Ativo': 'Ativu', 'ativa': 'Ativu', 'active': 'Ativu',
    'Paradu': 'Parado', 'Parado': 'Parado', 'inactive': 'Parado',
}

CREDIT_REPAYMENT_MAP = {
    'Selu diak hela': 'OnTime', 'Selu diak': 'OnTime', 'on time': 'OnTime',
    'Selu tarde': 'Late', 'late': 'Late', 'selu tarde': 'Late',
    'Banku macet': 'Stuck', 'stuck': 'Stuck', 'macet': 'Stuck',
    'Selu hotu ona': 'Done', 'done': 'Done', 'selu hotu': 'Done',
    'Falla': 'Failed', 'failed': 'Failed', 'falla': 'Failed',
}

def _clean(val):
    if val is None:
        return None
    return str(val).strip() if str(val).strip() else None

def _safe_int(val, default=0):
    try:
        if isinstance(val, str) and val.strip():
            return int(float(val))
        return int(val) if val not in [None, ''] else default
    except (ValueError, TypeError):
        return default

def _safe_float(val):
    try:
        if isinstance(val, str) and val.strip():
            return float(val.replace(',', ''))
        return float(val) if val not in [None, ''] else None
    except (ValueError, TypeError):
        return None

def _normalize_municipality(raw):
    if not raw:
        return None
    key = str(raw).strip().lower()
    return MUNICIPALITY_ALIASES.get(key, key)


# ══════════════════════════════════════════════════════════════════
# MAIN IMPORTER CLASS
# ══════════════════════════════════════════════════════════════════

class KredituSuaveImporter:
    def __init__(self, excel_file, user):
        self.excel_file = excel_file
        self.user = user
        self.errors = []
        self.warnings = []
        self.success_count = 0
        self.skipped_count = 0
        
        # Caches
        self.municipality_cache = {}
        self.admin_post_cache = {}
        self.village_cache = {}
        self.sector_cache = {}
        self.year_cache = {}
        self.faze_cache = {}
        self.status_cache = {}
        self.program_type_cache = {}
        
    def load_caches(self):
        """Load all reference data into caches"""
        for m in Municipality.objects.all():
            self.municipality_cache[m.name.lower()] = m
        for ap in AdministrativePost.objects.all():
            key = f"{ap.municipality_id}_{ap.name.lower()}"
            self.admin_post_cache[key] = ap
        for v in Village.objects.all():
            self.village_cache[v.name.lower()] = v
        for s in Sector.objects.all():
            self.sector_cache[s.name.lower()] = s
        for y in Year.objects.all():
            self.year_cache[y.year] = y
        for f in Faze.objects.all():
            self.faze_cache[f.name.upper()] = f
        for st in Status.objects.all():
            self.status_cache[st.name] = st
        
        # Program type for Kreditu Suave
        self.program_type_ks = TIpu_Programa.objects.filter(name='KREDITU SUAVE').first()
        if not self.program_type_ks:
            self.program_type_ks = TIpu_Programa.objects.create(name='KREDITU SUAVE')
        
        # Faze for Kreditu Suave
        self.faze_ks = self.faze_cache.get('KREDITU')
        if not self.faze_ks:
            self.faze_ks = Faze.objects.create(name='KREDITU')
    
    def get_or_create_benefisiariu(self, name, phone, gender, nivel_edukasaun=None):
        """Get or create Benefisiariu"""
        if not name:
            return None
        
        # Search by name or phone
        benef = None
        if phone:
            benef = Benefisiariu.objects.filter(phone=phone).first()
        if not benef and name:
            benef = Benefisiariu.objects.filter(name__iexact=name).first()
        
        if benef:
            # Update if needed
            if gender and not benef.sex:
                benef.sex = gender
                benef.save()
            return benef
        
        # Create new
        status = self.status_cache.get('Ativu', Status.objects.first())
        benef = Benefisiariu.objects.create(
            name=name[:100],
            phone=phone[:20] if phone else None,
            sex=gender if gender in ['Mane', 'Feto'] else None,
            nivel_edukasaun=nivel_edukasaun,
            status=status,
        )
        
        # Generate hashed after save (will be done in save method)
        return benef
    
    def create_address(self, benef, municipality, admin_post, village, aldeia, address):
        """Create AddressTL for benefisiariu"""
        AddressTL.objects.update_or_create(
            benefisiariu=benef,
            defaults={
                'address': address or '',
                'municipality': municipality,
                'administrativepost': admin_post,
                'village': village,
                'aldeia': aldeia or '',
            }
        )
    
    def create_business(self, benef, name, sector, idea=None):
        """Create Business"""
        business = Business.objects.create(
            benefisiariu=benef,
            name=name[:100] if name else None,
            idea=idea[:100] if idea else None,
            sector=sector,
        )
        return business
    
    def create_program(self, benef, amount, year_obj):
        """Create Program for Kreditu Suave"""
        if not amount:
            return None
        
        status = self.status_cache.get('Ativu', Status.objects.first())
        program = Program.objects.create(
            benefisiariu=benef,
            program_type=self.program_type_ks,
            faze=self.faze_ks,
            year=year_obj,
            approved_amount=amount,
            amount=amount,
            status=status,
        )
        return program
    
    def create_employee(self, business, male, female):
        """Create Employee"""
        Employee.objects.create(
            business=business,
            male=male or 0,
            female=female or 0,
        )
    
    def create_finance(self, business, budget):
        """Create Finance"""
        if budget:
            Finance.objects.create(
                business=business,
                budget=budget,
            )
    
    def create_credit_info(self, business, row):
        """Create CreditInfo from Excel row"""
        took_credit = row.get('took_credit')
        if took_credit is None:
            return
        
        provider = row.get('credit_provider')
        amount = row.get('credit_amount')
        satisfied = row.get('satisfied')
        wants_more = row.get('wants_more')
        preferred = row.get('preferred_institution')
        repayment_raw = row.get('repayment_status')
        
        repayment = None
        if repayment_raw:
            for key, val in CREDIT_REPAYMENT_MAP.items():
                if key.lower() in str(repayment_raw).lower():
                    repayment = val
                    break
        
        CreditInfo.objects.update_or_create(
            business=business,
            defaults={
                'took_credit': took_credit,
                'provider': provider,
                'amount': amount,
                'satisfied': satisfied,
                'wants_more': wants_more,
                'preferred_institution': preferred,
                'repayment_status': repayment,
            }
        )
    
    def create_financial_assessment(self, business, row):
        """Create FinancialAssessment from Excel row"""
        monthly_revenue = row.get('monthly_revenue')
        annual_revenue = row.get('annual_revenue')
        projected_revenue = row.get('projected_revenue')
        pays_tax = row.get('pays_tax')
        total_assets = row.get('total_assets')
        accounting = row.get('accounting_book')
        
        if not any([monthly_revenue, annual_revenue, total_assets]):
            return
        
        inventory_method = row.get('inventory_method')
        if inventory_method:
            inv_map = {'Manual': 'Manual', 'Software': 'Software', 'RFID': 'RFID'}
            inventory_method = inv_map.get(inventory_method, 'Manual')
        
        FinancialAssessment.objects.update_or_create(
            business=business,
            defaults={
                'monthly_revenue': monthly_revenue,
                'annual_revenue': annual_revenue,
                'projected_revenue': projected_revenue,
                'pays_tax': pays_tax,
                'total_assets': total_assets,
                'accounting_book': accounting,
                'inventory_method': inventory_method,
            }
        )
    
    def create_products(self, business, products_list):
        """Create ProductService from list of products"""
        for prod in products_list:
            if prod.get('name'):
                ProductService.objects.create(
                    business=business,
                    name=prod['name'][:200],
                    production_volume=str(prod.get('production_volume', '')),
                    production_frequency=prod.get('production_frequency', ''),
                    sales_volume=str(prod.get('sales_volume', '')),
                    sales_frequency=prod.get('sales_frequency', ''),
                    sales_amount=prod.get('sales_amount'),
                )
    
    def create_customers(self, business, customers_list):
        """Create MainCustomer from list"""
        for cust in customers_list:
            if cust.get('name'):
                MainCustomer.objects.create(
                    business=business,
                    name=cust['name'][:200],
                    demand_volume=str(cust.get('demand_volume', '')),
                    frequency=cust.get('frequency', ''),
                )
    
    def create_competitors(self, business, competitors_list):
        """Create Competitor from list"""
        for comp in competitors_list:
            if comp.get('name'):
                Competitor.objects.create(
                    business=business,
                    name=comp['name'][:200],
                    demand_volume=str(comp.get('demand_volume', '')),
                    frequency=comp.get('frequency', ''),
                )
    
    def create_market_assessment(self, business, row):
        """Create MarketAssessment from Excel row"""
        promotion = row.get('promotion_strategy')
        challenges = row.get('current_challenges')
        long_challenges = row.get('long_term_challenges')
        priority = row.get('priority')
        response = row.get('response_strategy')
        
        if not any([promotion, challenges, long_challenges]):
            return
        
        priority_map = {'High': 'High', 'Medium': 'Medium', 'Low': 'Low',
                       'Prioridade a\'as': 'High', 'Prioridade Mediu': 'Medium',
                       'Prioridade Ki\'ik': 'Low'}
        priority = priority_map.get(priority, 'Medium')
        
        MarketAssessment.objects.update_or_create(
            business=business,
            defaults={
                'promotion_strategy': promotion,
                'current_challenges': challenges,
                'long_term_challenges': long_challenges,
                'priority': priority,
                'response_strategy': response,
            }
        )


# ══════════════════════════════════════════════════════════════════
# IMPORT FUNCTION - From Dados Kredit Suave Sheet
# ══════════════════════════════════════════════════════════════════

def import_kreditu_suave_from_excel(excel_file, user):
    """Main import function for Kreditu Suave"""
    
    importer = KredituSuaveImporter(excel_file, user)
    importer.load_caches()
    
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    
    # Try to find the main data sheet
    sheet_names = wb.sheetnames
    target_sheet = None
    
    for sheet in sheet_names:
        if 'Dados Kredit Suave' in sheet or 'Monitorin' in sheet or 'Lista AProvadu' in sheet:
            target_sheet = sheet
            break
    
    if not target_sheet:
        return {'success': 0, 'errors': ['No valid sheet found']}
    
    ws = wb[target_sheet]
    
    # Find header row (around row 10-15)
    header_row = None
    for row in range(1, 20):
        first_cell = ws.cell(row=row, column=1).value
        if first_cell and 'No' in str(first_cell):
            header_row = row
            break
        if first_cell and 'Nu.' in str(first_cell):
            header_row = row
            break
    
    if not header_row:
        header_row = 11
    
    # Map columns
    col_map = {}
    for col in range(1, 30):
        val = ws.cell(row=header_row, column=col).value
        if val:
            val_str = str(val).strip().lower()
            if 'naran empreza' in val_str or '1.naran empreza' in val_str:
                col_map['naran_empreza'] = col
            elif 'naran emprezariu' in val_str or '2.naran emprezariu' in val_str:
                col_map['naran_emprezariu'] = col
            elif 'jeneru' in val_str or '3.jeneru' in val_str:
                col_map['jeneru'] = col
            elif 'nu. kontakto' in val_str or '6.nu. kontakto' in val_str:
                col_map['nu_kontaktu'] = col
            elif 'setor principal' in val_str or '8.setor principal' in val_str:
                col_map['setor'] = col
            elif 'municipio' in val_str:
                col_map['municipio'] = col
            elif 'postoadm' in val_str or 'posto adm' in val_str:
                col_map['posto'] = col
            elif 'suco' in val_str:
                col_map['suco'] = col
            elif 'aldeia' in val_str:
                col_map['aldeia'] = col
            elif 'total fundo aprovadu' in val_str or '13.total fundo' in val_str:
                col_map['total_fundo'] = col
            elif 'mane' in val_str and col_map.get('mane') is None:
                col_map['mane'] = col
            elif 'feto' in val_str and col_map.get('feto') is None:
                col_map['feto'] = col
            elif 'atividade negosiu' in val_str or '7.atividade' in val_str:
                col_map['atividade'] = col
    
    # Process data rows
    for row_num in range(header_row + 1, ws.max_row + 1):
        try:
            naran_empreza = _clean(ws.cell(row=row_num, column=col_map.get('naran_empreza')).value)
            naran_emprezariu = _clean(ws.cell(row=row_num, column=col_map.get('naran_emprezariu')).value)
            
            if not naran_empreza and not naran_emprezariu:
                continue
            
            # Get basic info
            name = naran_emprezariu or naran_empreza
            phone = _clean(ws.cell(row=row_num, column=col_map.get('nu_kontaktu')).value)
            gender = _clean(ws.cell(row=row_num, column=col_map.get('jeneru')).value)
            if gender:
                gender = 'Mane' if gender.lower() in ['m', 'mane'] else 'Feto'
            
            # Get or create benefisiariu
            benef = importer.get_or_create_benefisiariu(name, phone, gender)
            if not benef:
                importer.errors.append(f"Row {row_num}: Cannot create benefisiariu")
                continue
            
            # Municipality
            mun_name = _clean(ws.cell(row=row_num, column=col_map.get('municipio')).value)
            mun_normalized = _normalize_municipality(mun_name)
            mun_obj = importer.municipality_cache.get(mun_normalized)
            
            # Admin post
            post_name = _clean(ws.cell(row=row_num, column=col_map.get('posto')).value)
            post_obj = None
            if mun_obj and post_name:
                key = f"{mun_obj.id}_{post_name.lower()}"
                post_obj = importer.admin_post_cache.get(key)
            
            # Village
            village_name = _clean(ws.cell(row=row_num, column=col_map.get('suco')).value)
            village_obj = importer.village_cache.get(village_name.lower()) if village_name else None
            
            # Aldeia
            aldeia = _clean(ws.cell(row=row_num, column=col_map.get('aldeia')).value)
            
            # Create address
            importer.create_address(benef, mun_obj, post_obj, village_obj, aldeia, None)
            
            # LocBussiness (same as address)
            LocBussiness.objects.update_or_create(
                benefisiariu=benef,
                defaults={
                    'municipality': mun_obj,
                    'administrativepost': post_obj,
                    'village': village_obj,
                    'aldeia': aldeia,
                }
            )
            
            # Sector
            sector_name = _clean(ws.cell(row=row_num, column=col_map.get('setor')).value)
            sector_normalized = SECTOR_ALIASES.get(sector_name.lower(), sector_name) if sector_name else None
            sector_obj = importer.sector_cache.get(sector_normalized)
            
            # Business
            business = importer.create_business(benef, naran_empreza or name, sector_obj)
            
            # Employee
            male = _safe_int(ws.cell(row=row_num, column=col_map.get('mane')).value)
            female = _safe_int(ws.cell(row=row_num, column=col_map.get('feto')).value)
            importer.create_employee(business, male, female)
            
            # Program (Kreditu Suave)
            amount = _safe_float(ws.cell(row=row_num, column=col_map.get('total_fundo')).value)
            # Try to get year from sheet name or default
            year_match = re.search(r'(\d{4})', target_sheet)
            year_val = int(year_match.group(1)) if year_match else 2024
            year_obj = importer.year_cache.get(year_val)
            
            importer.create_program(benef, amount, year_obj)
            
            # Finance
            importer.create_finance(business, amount)
            
            importer.success_count += 1
            
        except Exception as e:
            importer.errors.append(f"Row {row_num}: {str(e)}")
    
    return {
        'success': importer.success_count,
        'errors': importer.errors,
        'warnings': importer.warnings
    }


# ══════════════════════════════════════════════════════════════════
# DJANGO VIEWS
# ══════════════════════════════════════════════════════════════════

from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from config.decorators import allowed_users

@login_required
@allowed_users(allowed_roles=['admin', 'KS'])
def import_kreditu_suave_excel(request):
    """Import Kreditu Suave data from Excel"""
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, "Formatu arquivo tenke .xlsx ka .xls!")
            return redirect('import_ks_excel')
        
        result = import_kreditu_suave_from_excel(excel_file, request.user)
        
        if result['success']:
            messages.success(request, f"✅ Suksesu importa {result['success']} dadus Kreditu Suave!")
        
        for error in result['errors'][:10]:
            messages.warning(request, f"⚠️ {error}")
        
        return redirect('dash-ks')
    
    context = {
        'title': 'Import Dadus Excel Kreditu Suave',
        'legend': 'Import Dadus Excel Kreditu Suave',
    }
    return render(request, 'Dash_ks/import.html', context)


@login_required
@allowed_users(allowed_roles=['admin', 'KS'])
def export_kreditu_suave_excel(request):
    """Export Kreditu Suave data to Excel"""
    from openpyxl import Workbook
    from django.http import HttpResponse
    
    wb = Workbook()
    
    # Sheet 1: Benefisiariu & Business
    ws1 = wb.active
    ws1.title = "Benefisiariu KS"
    ws1.append(['No', 'Naran Benefisiariu', 'Sexo', 'Telefone', 'Naran Empreza', 
                'Setor', 'Munisipiu', 'Montante Kreditu', 'Tinan', 'Status'])
    
    programs = Program.objects.filter(program_type__name='KREDITU SUAVE').select_related(
        'benefisiariu', 'year', 'status'
    )
    
    for i, prog in enumerate(programs, 1):
        benef = prog.benefisiariu
        business = Business.objects.filter(benefisiariu=benef).first()
        loc = LocBussiness.objects.filter(benefisiariu=benef).first()
        
        ws1.append([
            i,
            benef.name if benef else '',
            benef.sex or '',
            benef.phone or '',
            business.name if business else '',
            str(business.sector) if business and business.sector else '',
            str(loc.municipality) if loc and loc.municipality else '',
            float(prog.amount) if prog.amount else 0,
            prog.year.year if prog.year else '',
            str(prog.status) if prog.status else ''
        ])
    
    # Sheet 2: Credit Info
    ws2 = wb.create_sheet("Informasaun Kreditu")
    ws2.append(['Empreza', 'Foti Kreditu Tan?', 'Provider', 'Amount', 'Repayment Status'])
    
    for credit in CreditInfo.objects.select_related('business'):
        ws2.append([
            str(credit.business),
            'Sim' if credit.took_credit else 'Lae',
            credit.provider or '',
            float(credit.amount) if credit.amount else 0,
            credit.get_repayment_status_display() or ''
        ])
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="kreditu_suave_export.xlsx"'
    wb.save(response)
    return response