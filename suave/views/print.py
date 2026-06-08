import io
from datetime import date

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render

from config.decorators import allowed_users
from custom.models import Municipality, Year, Faze, Sector
from benefisiariu.models import Benefisiariu
from kni.models import Business, LocBussiness, Program, Employee, Finance
from suave.models import (
    EkipaMember, ProductService, MainCustomer, Competitor,
    MarketAssessment, FinancialAssessment, CreditInfo,
)


# ---------------------------------------------------------------------------
# Helper: queryset Benefisiariu Kreditu Suave + filter
# ---------------------------------------------------------------------------

def _get_filtered_qs(params):
    qs = (
        Benefisiariu.active_objects
        .filter(Pnegosiu__program_type__name="KREDITU SUAVE")
        .select_related("status")
        .prefetch_related(
            "locnegosiu__municipality",
            "locnegosiu__administrativepost",
            "locnegosiu__village",
            "Pnegosiu__program_type",
            "Pnegosiu__faze",
            "Pnegosiu__year",
            "Pnegosiu__status",
            "Pnegosiu__t_apoiu",
            "Pnegosiu__t_fundus",
            "negosiu__sector",
            "negosiu__category",
            "negosiu__size",
            "negosiu__employee_set",        # related_name default Django
            "negosiu__finance_set",         # related_name default Django
            "negosiu__products",            # ProductService.related_name
            "negosiu__customers",           # MainCustomer.related_name
            "negosiu__competitors",         # Competitor.related_name
            "negosiu__market_assessment",   # MarketAssessment (OneToOne)
            "negosiu__financial_assessment",# FinancialAssessment (OneToOne)
            "negosiu__credit_info",         # CreditInfo (OneToOne)
            "team_members",                 # EkipaMember
        )
        .order_by("locnegosiu__municipality__name", "name")
        .distinct()
    )

    mun       = params.get("mun",       "").strip()
    year      = params.get("year",      "").strip()
    sector    = params.get("sector",    "").strip()
    sexo      = params.get("sexo",      "").strip()
    status    = params.get("status",    "").strip()
    repayment = params.get("repayment", "").strip()

    if mun:       qs = qs.filter(locnegosiu__municipality__name__icontains=mun)
    if year:      qs = qs.filter(Pnegosiu__year__year=year)
    if sector:    qs = qs.filter(negosiu__sector__name__icontains=sector)
    if sexo:      qs = qs.filter(sex__iexact=sexo)
    if status:    qs = qs.filter(Pnegosiu__status__name__iexact=status)
    if repayment: qs = qs.filter(negosiu__credit_info__repayment_status=repayment)

    return qs


def _filter_label(params):
    parts = []
    if params.get("mun"):       parts.append(f"Munisipiu: {params['mun']}")
    if params.get("year"):      parts.append(f"Tinan: {params['year']}")
    if params.get("sector"):    parts.append(f"Setor: {params['sector']}")
    if params.get("sexo"):      parts.append(f"Genero: {params['sexo']}")
    if params.get("status"):    parts.append(f"Status: {params['status']}")
    if params.get("repayment"): parts.append(f"Pagamentu: {params['repayment']}")
    return " | ".join(parts) if parts else "Dadus Tomak Kreditu Suave (Hotu)"


# ---------------------------------------------------------------------------
# Accessor helpers
# ---------------------------------------------------------------------------

def _biz(b):    return b.negosiu.first()
def _loc(b):    return b.locnegosiu.first()
def _prog(b):   return b.Pnegosiu.filter(program_type__name="KREDITU SUAVE").first()
def _emp(b):
    biz = _biz(b)
    return biz.employee_set.first() if biz else None
def _fin(b):
    biz = _biz(b)
    return biz.finance_set.first() if biz else None
def _mkt(b):
    biz = _biz(b)
    try:    return biz.market_assessment if biz else None
    except: return None
def _fasmt(b):
    biz = _biz(b)
    try:    return biz.financial_assessment if biz else None
    except: return None
def _credit(b):
    biz = _biz(b)
    try:    return biz.credit_info if biz else None
    except: return None
def _xefi(b):
    return b.team_members.filter(role="Xefi").first()
def _tekniko(b):
    return b.team_members.filter(role="Tekniko").first()

def _safe(fn):
    try:
        v = fn()
        return v if v is not None else ""
    except:
        return ""


# ---------------------------------------------------------------------------
# Definisi kolom — 4 sheet sesuai format file Excel
# ---------------------------------------------------------------------------

# Sheet 1: Dados Kredit Suave (monitoring lapangan)
DADOS_HEADERS = [
    ("No",                      None),
    ("Xefi Ekipa",              lambda b: _safe(lambda: _xefi(b).name if _xefi(b) else "")),
    ("Nu. Kontakto Xefi",       lambda b: _safe(lambda: _xefi(b).phone or "" if _xefi(b) else "")),
    ("Tekniko",                 lambda b: _safe(lambda: _tekniko(b).name if _tekniko(b) else "")),
    ("Nu. Kontakto Tekniko",    lambda b: _safe(lambda: _tekniko(b).phone or "" if _tekniko(b) else "")),
    ("Naran Empreza",           lambda b: _safe(lambda: _biz(b).name or "")),
    ("Naran Emprezariu",        lambda b: b.name or ""),
    ("Jeneru",                  lambda b: b.sex or ""),
    ("Nivel Edukasaun",         lambda b: b.get_nivel_edukasaun_display() if b.nivel_edukasaun else ""),
    ("Nu. Kontakto",            lambda b: b.phone or ""),
    ("Email / Website",         lambda b: b.email_website or ""),
    ("Atividade Negosiu",       lambda b: _safe(lambda: _biz(b).idea or "")),
    ("Setor Prinsipal",         lambda b: _safe(lambda: str(_biz(b).sector) if _biz(b) and _biz(b).sector else "")),
    ("Kategoria Empreza",       lambda b: _safe(lambda: str(_biz(b).category) if _biz(b) and _biz(b).category else "")),
    ("Status Atividade",        lambda b: _safe(lambda: str(_prog(b).status) if _prog(b) and _prog(b).status else "")),
    ("Lokalizasaun Negosiu",    lambda b: _safe(lambda: f"{_loc(b).municipality}, {_loc(b).administrativepost}" if _loc(b) and _loc(b).municipality else "")),
    ("Municipio",               lambda b: _safe(lambda: str(_loc(b).municipality) if _loc(b) and _loc(b).municipality else "")),
    ("Posto Administrativo",    lambda b: _safe(lambda: str(_loc(b).administrativepost) if _loc(b) and _loc(b).administrativepost else "")),
    ("Suco",                    lambda b: _safe(lambda: str(_loc(b).village) if _loc(b) and _loc(b).village else "")),
    ("Aldeia",                  lambda b: _safe(lambda: _loc(b).aldeia or "" if _loc(b) else "")),
    ("Total Fundo Aprovadu ($)",lambda b: _safe(lambda: float(_prog(b).approved_amount or 0) if _prog(b) else 0)),
    ("Nu. Trabalhador Atual",   lambda b: _safe(lambda: _emp(b).total if _emp(b) else 0)),
    ("Feto",                    lambda b: _safe(lambda: _emp(b).female if _emp(b) else 0)),
    ("Mane",                    lambda b: _safe(lambda: _emp(b).male if _emp(b) else 0)),
    ("Amount Kreditu ($)",      lambda b: _safe(lambda: float(_credit(b).amount or 0) if _credit(b) else 0)),
    ("Implicasaun Renebus",     lambda b: _safe(lambda: _credit(b).get_repayment_status_display() if _credit(b) and _credit(b).repayment_status else "")),
    ("Rekomendasaun",           lambda b: _safe(lambda: _credit(b).recommendation or "" if _credit(b) else "")),
    ("Observasaun",             lambda b: _safe(lambda: _credit(b).repayment_notes or "" if _credit(b) else "")),
    ("Tinan",                   lambda b: _safe(lambda: str(_prog(b).year) if _prog(b) and _prog(b).year else "")),
]

# Sheet 2: Informasaun FGKS (data bank)
FGKS_HEADERS = [
    ("No",                          None),
    ("Entrepreneur's Name",         lambda b: b.name or ""),
    ("Gender",                      lambda b: b.sex or ""),
    ("Enterprise Name",             lambda b: _safe(lambda: _biz(b).name or "")),
    ("Sectors",                     lambda b: _safe(lambda: str(_biz(b).sector) if _biz(b) and _biz(b).sector else "")),
    ("Business Activities",         lambda b: _safe(lambda: _biz(b).idea or "")),
    ("Municipality",                lambda b: _safe(lambda: str(_loc(b).municipality) if _loc(b) and _loc(b).municipality else "")),
    ("Approved by Bank",            lambda b: _safe(lambda: _credit(b).approved_by_bank or "" if _credit(b) else "")),
    ("Amount Approved ($)",         lambda b: _safe(lambda: float(_prog(b).approved_amount or 0) if _prog(b) else 0)),
    ("Amount of Collateral ($)",    lambda b: _safe(lambda: float(_credit(b).collateral_amount or 0) if _credit(b) else 0)),
    ("Employees (Total)",           lambda b: _safe(lambda: _emp(b).total if _emp(b) else 0)),
    ("Female",                      lambda b: _safe(lambda: _emp(b).female if _emp(b) else 0)),
    ("Male",                        lambda b: _safe(lambda: _emp(b).male if _emp(b) else 0)),
    ("Annual Revenue ($)",          lambda b: _safe(lambda: float(_fasmt(b).annual_revenue or 0) if _fasmt(b) else 0)),
    ("Total Assets ($)",            lambda b: _safe(lambda: float(_fasmt(b).total_assets or 0) if _fasmt(b) else 0)),
    ("Business Size",               lambda b: _safe(lambda: str(_biz(b).size) if _biz(b) and _biz(b).size else "")),
    ("Bookeeping",                  lambda b: _safe(lambda: _fasmt(b).get_accounting_book_display() if _fasmt(b) and _fasmt(b).accounting_book else "")),
    ("Business Challenges",         lambda b: _safe(lambda: _mkt(b).current_challenges or "" if _mkt(b) else "")),
    ("Any problem repaying loans",  lambda b: _safe(lambda: "Iha" if _credit(b) and _credit(b).has_repayment_problem else "Laiha")),
    ("Programs to be Continued",    lambda b: _safe(lambda: "Sin" if _credit(b) and _credit(b).program_continuation else "Lae")),
    ("Access Again",                lambda b: _safe(lambda: "Sin" if _credit(b) and _credit(b).wants_more else "Lae")),
    ("Observation",                 lambda b: _safe(lambda: _credit(b).repayment_notes or "" if _credit(b) else "")),
    ("Tinan",                       lambda b: _safe(lambda: str(_prog(b).year) if _prog(b) and _prog(b).year else "")),
]

# Sheet 3: Avaliasaun Merkadu
MARKET_HEADERS = [
    ("No",                          None),
    ("Naran Emprezariu",            lambda b: b.name or ""),
    ("Naran Empreza",               lambda b: _safe(lambda: _biz(b).name or "")),
    ("Setor",                       lambda b: _safe(lambda: str(_biz(b).sector) if _biz(b) and _biz(b).sector else "")),
    ("Municipio",                   lambda b: _safe(lambda: str(_loc(b).municipality) if _loc(b) and _loc(b).municipality else "")),
    ("Estratejia Promosaun",        lambda b: _safe(lambda: _mkt(b).promotion_strategy or "" if _mkt(b) else "")),
    ("Difikuldade Atual",           lambda b: _safe(lambda: _mkt(b).current_challenges or "" if _mkt(b) else "")),
    ("Dezafio Longu Prazu",         lambda b: _safe(lambda: _mkt(b).long_term_challenges or "" if _mkt(b) else "")),
    ("Prioridade Dezafio",          lambda b: _safe(lambda: _mkt(b).get_priority_display() if _mkt(b) and _mkt(b).priority else "")),
    ("Estratejia Responde Dezafio", lambda b: _safe(lambda: _mkt(b).response_strategy or "" if _mkt(b) else "")),
    ("Kliente Prinsipal",           lambda b: _safe(lambda: ", ".join(c.name for c in _biz(b).customers.all()[:3]) if _biz(b) else "")),
    ("Kompetitor Prinsipal",        lambda b: _safe(lambda: ", ".join(c.name for c in _biz(b).competitors.all()[:3]) if _biz(b) else "")),
]

# Sheet 4: Avaliasaun Finanseiru + Kreditu
FINANCIAL_HEADERS = [
    ("No",                          None),
    ("Naran Emprezariu",            lambda b: b.name or ""),
    ("Naran Empreza",               lambda b: _safe(lambda: _biz(b).name or "")),
    ("Livru Kontabilidade",         lambda b: _safe(lambda: _fasmt(b).get_accounting_book_display() if _fasmt(b) and _fasmt(b).accounting_book else "")),
    ("Metodu Inventariu",           lambda b: _safe(lambda: _fasmt(b).get_inventory_method_display() if _fasmt(b) and _fasmt(b).inventory_method else "")),
    ("Rendimentu Fulan ($)",        lambda b: _safe(lambda: float(_fasmt(b).monthly_revenue or 0) if _fasmt(b) else 0)),
    ("Rendimentu Tinan ($)",        lambda b: _safe(lambda: float(_fasmt(b).annual_revenue or 0) if _fasmt(b) else 0)),
    ("Projeksaun Rendimentu ($)",   lambda b: _safe(lambda: float(_fasmt(b).projected_revenue or 0) if _fasmt(b) else 0)),
    ("Selu Taxa?",                  lambda b: _safe(lambda: "Sin" if _fasmt(b) and _fasmt(b).pays_tax else "Lae")),
    ("Taxa Fulan ($)",              lambda b: _safe(lambda: float(_fasmt(b).monthly_tax or 0) if _fasmt(b) else 0)),
    ("Total Assets ($)",            lambda b: _safe(lambda: float(_fasmt(b).total_assets or 0) if _fasmt(b) else 0)),
    ("Foti Tan Kreditu?",           lambda b: _safe(lambda: "Sin" if _credit(b) and _credit(b).took_credit else "Lae")),
    ("Provider Kreditu",            lambda b: _safe(lambda: _credit(b).provider or "" if _credit(b) else "")),
    ("Montante Kreditu ($)",        lambda b: _safe(lambda: float(_credit(b).amount or 0) if _credit(b) else 0)),
    ("Kolateral ($)",               lambda b: _safe(lambda: float(_credit(b).collateral_amount or 0) if _credit(b) else 0)),
    ("Tuir Espetativa?",            lambda b: _safe(lambda: "Sin" if _credit(b) and _credit(b).satisfied else "Lae")),
    ("Status Pagamentu",            lambda b: _safe(lambda: _credit(b).get_repayment_status_display() if _credit(b) and _credit(b).repayment_status else "")),
    ("Iha Problema Pagamentu?",     lambda b: _safe(lambda: "Sin" if _credit(b) and _credit(b).has_repayment_problem else "Lae")),
    ("Hakarak Kreditu Futuro?",     lambda b: _safe(lambda: "Sin" if _credit(b) and _credit(b).wants_more else "Lae")),
    ("Institusaun Preferida",       lambda b: _safe(lambda: _credit(b).preferred_institution or "" if _credit(b) else "")),
    ("Programa Kontinua?",          lambda b: _safe(lambda: "Sin" if _credit(b) and _credit(b).program_continuation else "Lae")),
    ("Rekomendasaun",               lambda b: _safe(lambda: _credit(b).recommendation or "" if _credit(b) else "")),
    ("Aprovasaun Banku",            lambda b: _safe(lambda: _credit(b).approved_by_bank or "" if _credit(b) else "")),
]


# ===========================================================================
# EXPORT EXCEL — 4 sheets
# ===========================================================================

@login_required
@allowed_users(allowed_roles=["admin", "KS", "XFD"])
def export_excel_ks(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    params     = request.GET
    qs         = _get_filtered_qs(params)
    benef_list = list(qs)
    label      = _filter_label(params)
    today      = date.today().strftime("%d-%m-%Y")
    filename   = f"KreditoSuave_Export_{today}.xlsx"

    wb = Workbook()

    # ── Style constants ──
    GREEN  = "1A5276"
    BLUE   = "1F3864"
    PURPLE = "4A235A"
    ORANGE = "7E5109"
    WHITE  = "FFFFFF"
    GRAY   = "F2F2F2"
    LBLUE  = "D6E4F0"

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin   = Border(
        left=Side(style="thin"),  right=Side(style="thin"),
        top=Side(style="thin"),   bottom=Side(style="thin"),
    )

    def fill(h): return PatternFill("solid", fgColor=h)

    def _write_sheet(ws, headers, data, title_text, hdr_hex):
        # Baris 1: judul — TIDAK merge_cells (avoid MergedCell error)
        ws["A1"] = title_text
        ws["A1"].font      = Font(name="Arial", bold=True, size=12, color=hdr_hex)
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 22

        # Baris 2: filter & tanggal
        ws["A2"] = f"Filter: {label}  |  Rai tiha: {today}"
        ws["A2"].font      = Font(name="Arial", size=9, italic=True, color="555555")
        ws["A2"].alignment = center
        ws.row_dimensions[2].height = 16

        ws.row_dimensions[3].height = 6  # spacer

        # Baris 4: header kolom
        hfill = fill(hdr_hex)
        hfont = Font(name="Arial", bold=True, color=WHITE, size=9)
        for col, (name, _) in enumerate(headers, 1):
            c = ws.cell(row=4, column=col, value=name)
            c.font = hfont; c.fill = hfill
            c.alignment = center; c.border = thin
        ws.row_dimensions[4].height = 32

        # Baris 5+: data
        DATA_START = 5
        nfont = Font(name="Arial", size=9)
        for i, benef in enumerate(data, 1):
            rfill = fill(GRAY) if i % 2 == 0 else fill(WHITE)
            for col, (_, acc) in enumerate(headers, 1):
                val = i if acc is None else _safe(lambda a=acc, b=benef: a(b))
                c = ws.cell(row=DATA_START + i - 1, column=col, value=val)
                c.font = nfont; c.fill = rfill
                c.border = thin
                c.alignment = center if col == 1 else left
            ws.row_dimensions[DATA_START + i - 1].height = 16

        # Baris total — per sel, TIDAK merge
        total_row = DATA_START + len(data)
        tfont = Font(name="Arial", bold=True, size=9, color=hdr_hex)
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=total_row, column=col, value="")
            c.fill = fill(LBLUE); c.border = thin
        ws.cell(row=total_row, column=1,
                value=f"Total Rekord: {len(data)}").font = tfont

        # Lebar kolom
        for col, (name, _) in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col)].width = max(10, min(len(name) + 4, 30))

        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}4"

    # ── Sheet 1: Dados Kredit Suave ──
    ws1 = wb.active
    ws1.title = "Dados Kredit Suave"
    _write_sheet(ws1, DADOS_HEADERS, benef_list,
                 "REKAPITULASAUN DADOS KREDIT SUAVE – MONITORIZASAUN LAPANGAN",
                 GREEN)

    # ── Sheet 2: Informasaun FGKS ──
    ws2 = wb.create_sheet("Informasaun FGKS")
    _write_sheet(ws2, FGKS_HEADERS, benef_list,
                 "INFORMASAUN FGKS – FASILIDADE GARANTIA KREDITU SUAVE",
                 BLUE)

    # ── Sheet 3: Avaliasaun Merkadu ──
    ws3 = wb.create_sheet("Avaliasaun Merkadu")
    with_mkt = [b for b in benef_list if _mkt(b)]
    _write_sheet(ws3, MARKET_HEADERS, with_mkt,
                 "AVALIASAUN MERKADU – ESTRATEJIA NO DEZAFIU",
                 PURPLE)

    # ── Sheet 4: Avaliasaun Finanseiru + Kreditu ──
    ws4 = wb.create_sheet("Avaliasaun Finanseiru")
    _write_sheet(ws4, FINANCIAL_HEADERS, benef_list,
                 "AVALIASAUN FINANSEIRU NO INFORMASAUN KREDITU",
                 ORANGE)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ===========================================================================
# EXPORT PDF — landscape A3, kolom ringkas sheet 1
# ===========================================================================

@login_required
@allowed_users(allowed_roles=["admin", "KS", "XFD"])
def export_pdf_ks(request):
    from reportlab.lib           import colors
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib.styles    import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units     import cm
    from reportlab.platypus      import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, HRFlowable,
    )

    params     = request.GET
    qs         = _get_filtered_qs(params)
    benef_list = list(qs)
    label      = _filter_label(params)
    today      = date.today().strftime("%d/%m/%Y")
    filename   = f"KreditoSuave_Export_{date.today().strftime('%d-%m-%Y')}.pdf"

    PDF_COLS = [
        ("No",              None),
        ("Naran Emprezariu",lambda b: b.name or ""),
        ("Jeneru",          lambda b: b.sex or ""),
        ("Naran Empreza",   lambda b: _safe(lambda: _biz(b).name or "")),
        ("Setor",           lambda b: _safe(lambda: str(_biz(b).sector) if _biz(b) and _biz(b).sector else "")),
        ("Atividade",       lambda b: _safe(lambda: _biz(b).idea or "")),
        ("Mane",            lambda b: _safe(lambda: _emp(b).male if _emp(b) else 0)),
        ("Feto",            lambda b: _safe(lambda: _emp(b).female if _emp(b) else 0)),
        ("Total Trab.",     lambda b: _safe(lambda: _emp(b).total if _emp(b) else 0)),
        ("Municipio",       lambda b: _safe(lambda: str(_loc(b).municipality) if _loc(b) and _loc(b).municipality else "")),
        ("Posto",           lambda b: _safe(lambda: str(_loc(b).administrativepost) if _loc(b) and _loc(b).administrativepost else "")),
        ("Tinan",           lambda b: _safe(lambda: str(_prog(b).year) if _prog(b) and _prog(b).year else "")),
        ("Status",          lambda b: _safe(lambda: str(_prog(b).status) if _prog(b) and _prog(b).status else "")),
        ("Fundo Aprova ($)",lambda b: _safe(lambda: float(_prog(b).approved_amount or 0) if _prog(b) else 0)),
        ("Status Pagam.",   lambda b: _safe(lambda: _credit(b).get_repayment_status_display() if _credit(b) and _credit(b).repayment_status else "")),
    ]

    buffer = io.BytesIO()
    page_w, _ = landscape(A3)
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A3),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm,  bottomMargin=1.5*cm,
        title="Dados Kreditu Suave",
    )

    styles     = getSampleStyleSheet()
    GREEN      = colors.HexColor("#1A5276")
    GRAY       = colors.HexColor("#F2F2F2")
    cell_style = ParagraphStyle("Cell", parent=styles["Normal"], fontSize=7, leading=9)

    total_emp  = sum(_safe(lambda b=b: _emp(b).total if _emp(b) else 0) or 0 for b in benef_list)
    total_fund = sum(_safe(lambda b=b: float(_prog(b).approved_amount or 0) if _prog(b) else 0) or 0 for b in benef_list)

    story = [
        Paragraph(
            "REKAPITULASAUN DADOS KREDIT SUAVE<br/>FASILIDADE GARANTIA KREDITU SUAVE",
            ParagraphStyle("T", parent=styles["Heading1"], fontSize=13,
                           textColor=GREEN, spaceAfter=4, alignment=1),
        ),
        Paragraph(
            f"Filter: {label}  |  Rai tiha: {today}",
            ParagraphStyle("S", parent=styles["Normal"], fontSize=8,
                           textColor=colors.HexColor("#555555"),
                           spaceAfter=8, alignment=1),
        ),
        HRFlowable(width="100%", thickness=1, color=GREEN, spaceAfter=8),
    ]

    tbl_data = [[c[0] for c in PDF_COLS]]
    for i, benef in enumerate(benef_list, 1):
        row = []
        for col_name, acc in PDF_COLS:
            if acc is None:
                row.append(str(i))
            else:
                val = str(_safe(lambda a=acc, b=benef: a(b)))
                row.append(Paragraph(val, cell_style) if len(val) > 18 else val)
        tbl_data.append(row)

    usable_w   = page_w - 3 * cm
    col_ratios = [0.04, 0.11, 0.05, 0.11, 0.08, 0.10,
                  0.04, 0.04, 0.05, 0.08, 0.08, 0.05, 0.07, 0.08, 0.08]
    col_widths = [usable_w * r for r in col_ratios]

    row_bg = [
        ("BACKGROUND", (0, idx), (-1, idx),
         GRAY if idx % 2 == 0 else colors.white)
        for idx in range(1, len(tbl_data))
    ]

    tbl = Table(tbl_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), GREEN),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0), 7),
        ("ALIGN",         (0, 0), (-1, 0), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",      (0, 1), (-1, -1), 7),
        ("ALIGN",         (0, 1), (0, -1),  "CENTER"),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("LINEBELOW",     (0, 0), (-1, 0),  1,   GREEN),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING",   (0, 0), (-1, -1), 4),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
        *row_bg,
    ]))

    story += [
        tbl,
        Spacer(1, 0.4*cm),
        Paragraph(
            f"<b>Total Rekord: {len(benef_list)}</b> &nbsp;|&nbsp; "
            f"<b>Total Trabalhador: {total_emp}</b> &nbsp;|&nbsp; "
            f"<b>Total Fundo Aprovadu: ${total_fund:,.2f}</b> &nbsp;|&nbsp; "
            f"Imprimi tiha: {today}",
            ParagraphStyle("Footer", parent=styles["Normal"],
                           fontSize=8, textColor=GREEN),
        ),
    ]

    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ===========================================================================
# HALAMAN FILTER
# ===========================================================================

@login_required
@allowed_users(allowed_roles=["admin", "KS", "XFD"])
def export_page_ks(request):
    params = request.GET
    qs     = _get_filtered_qs(params)
    label  = _filter_label(params)

    context = {
        "title":   "Export Dados Kreditu Suave",
        "legend":  "Export Dados Kreditu Suave (Excel / PDF)",
        "group":   request.user.groups.all()[0].name,
        "total":   qs.count(),
        "label":   label,
        "params":  params.urlencode(),
        "mun_list":    Municipality.active_objects.order_by("name"),
        "year_list":   Year.active_objects.order_by("-year"),
        "sector_list": Sector.active_objects.order_by("name"),
        "repayment_choices": [
            ("OnTime", "Selu Diak Hela"),
            ("Late",   "Selu Tarde"),
            ("Stuck",  "Banku Macet"),
            ("Done",   "Selu Hotu Ona"),
            ("Failed", "Falla"),
            ("Other",  "Seluk"),
        ],
        "sel_mun":       params.get("mun",       ""),
        "sel_year":      params.get("year",       ""),
        "sel_sector":    params.get("sector",     ""),
        "sel_sexo":      params.get("sexo",       ""),
        "sel_status":    params.get("status",     ""),
        "sel_repayment": params.get("repayment",  ""),
        # Kolom untuk tampilan di template
        "excel_cols_sheet1": [h[0] for h in DADOS_HEADERS if h[0] != "No"],
        "excel_cols_sheet2": [h[0] for h in FGKS_HEADERS  if h[0] != "No"],
    }
    return render(request, "Print/export_ks.html", context)


