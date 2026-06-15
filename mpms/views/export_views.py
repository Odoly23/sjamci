import io
from datetime import date

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render

from config.decorators import allowed_users
from custom.models import Municipality
from mpms.models import (
    mpmsEmpresa, mpmsLokalizasaun, mpmsLisensamentu,
    mpmsKapital, mpmsEmpregador, mpmsMateriaPrima, mpmsAtividade,
    TIPO_ATIVIDADE,
)


# ---------------------------------------------------------------------------
# Helper: queryset dengan filter
# ---------------------------------------------------------------------------

def _get_filtered_qs(params):
    qs = (
        mpmsEmpresa.active_objects
        .select_related(
            "benefisiariu",
            "tipo_atividade",
            "lokalizasaun__municipality",
            "lokalizasaun__administrativepost",
            "lokalizasaun__village",
            "lisensamentu",
            "kapital",
            "empregador",
            "materia_prima",
        )
        .prefetch_related("atividades__year", "atividades__status")
        .order_by("lokalizasaun__municipality__name", "company_name")
    )

    mun          = params.get("mun", "").strip()
    tinan        = params.get("tinan", "").strip()
    tipo         = params.get("tipo", "").strip()
    sexo         = params.get("sexo", "").strip()
    lisensamentu = params.get("lisensamentu", "").strip()
    tipu_fundus  = params.get("tipu_fundus", "").strip()

    if mun:          qs = qs.filter(lokalizasaun__municipality__name__icontains=mun)
    if tinan:        qs = qs.filter(tinan_hari=tinan)
    if tipo:         qs = qs.filter(tipo_atividade__name__icontains=tipo)
    if sexo:         qs = qs.filter(benefisiariu__sex__iexact=sexo)
    if lisensamentu: qs = qs.filter(lisensamentu__lisensamentu__iexact=lisensamentu)
    if tipu_fundus:  qs = qs.filter(kapital__tipu_fundus__iexact=tipu_fundus)

    return qs


def _filter_label(params):
    parts = []
    if params.get("mun"):          parts.append(f"Munisipiu: {params['mun']}")
    if params.get("tinan"):        parts.append(f"Tinan: {params['tinan']}")
    if params.get("tipo"):         parts.append(f"Tipo: {params['tipo']}")
    if params.get("sexo"):         parts.append(f"Sexo: {params['sexo']}")
    if params.get("lisensamentu"): parts.append(f"Lisensamentu: {params['lisensamentu']}")
    if params.get("tipu_fundus"):  parts.append(f"Fundus: {params['tipu_fundus']}")
    return " | ".join(parts) if parts else "Dadus Tomak (Hotu)"


# ---------------------------------------------------------------------------
# Definisi kolom — dipakai bersama oleh Excel dan PDF
# ---------------------------------------------------------------------------

HEADERS = [
    ("No",                      None),
    ("Tipo Atividade",          lambda e: str(e.tipo_atividade) if e.tipo_atividade else ""),
    ("Naran Diretor",           lambda e: e.benefisiariu.name if e.benefisiariu else ""),
    ("Naran Kompania",          lambda e: e.company_name or ""),
    ("Sexo",                    lambda e: e.benefisiariu.sex if e.benefisiariu else ""),
    ("Nivel Edukasaun",         lambda e: getattr(e.benefisiariu, "nivel_edukasaun", "") or ""),
    ("Munisipiu",               lambda e: str(e.lokalizasaun.municipality) if getattr(e, "lokalizasaun", None) else ""),
    ("Postu Administrativu",    lambda e: str(e.lokalizasaun.administrativepost) if getattr(e, "lokalizasaun", None) else ""),
    ("Suku",                    lambda e: str(e.lokalizasaun.village) if getattr(e, "lokalizasaun", None) else ""),
    ("Aldeia",                  lambda e: (e.lokalizasaun.aldeia or "") if getattr(e, "lokalizasaun", None) else ""),
    ("Tinan Hari Kompania",     lambda e: e.tinan_hari or ""),
    ("Lisensamentu",            lambda e: e.lisensamentu.get_lisensamentu_display() if getattr(e, "lisensamentu", None) else ""),
    ("Status Lisensamentu",     lambda e: e.lisensamentu.get_lisensamentu_status_display() if getattr(e, "lisensamentu", None) else ""),
    ("Tipo Rai",                lambda e: e.lisensamentu.get_tipo_rai_display() if getattr(e, "lisensamentu", None) else ""),
    ("Kapital Investimento",    lambda e: e.kapital.get_kapital_investimento_display() if getattr(e, "kapital", None) else ""),
    ("Tipu Fundus",             lambda e: e.kapital.get_tipu_fundus_display() if getattr(e, "kapital", None) else ""),
    ("Total Fundus",            lambda e: e.kapital.get_total_fundus_display() if getattr(e, "kapital", None) else ""),
    ("Lukru Brutu/Mes ($)",     lambda e: e.kapital.lukru_brutu_mes or 0 if getattr(e, "kapital", None) else 0),
    ("Lukru Brutu/Ano ($)",     lambda e: e.kapital.lukru_brutu_ano or 0 if getattr(e, "kapital", None) else 0),
    ("Lukru Likidu/Mes ($)",    lambda e: e.kapital.lukru_likidu_mes or 0 if getattr(e, "kapital", None) else 0),
    ("Lukru Likidu/Ano ($)",    lambda e: e.kapital.lukru_likidu_ano or 0 if getattr(e, "kapital", None) else 0),
    ("Empregador Nas. Mane",    lambda e: e.empregador.nasional_mane if getattr(e, "empregador", None) else 0),
    ("Empregador Nas. Feto",    lambda e: e.empregador.nasional_feto if getattr(e, "empregador", None) else 0),
    ("Empregador Int. Mane",    lambda e: e.empregador.internasional_mane if getattr(e, "empregador", None) else 0),
    ("Empregador Int. Feto",    lambda e: e.empregador.internasional_feto if getattr(e, "empregador", None) else 0),
    ("Total Empregador",        lambda e: e.empregador.total_empregador if getattr(e, "empregador", None) else 0),
    ("Kustu Materia Prima ($)", lambda e: e.materia_prima.kustu or 0 if getattr(e, "materia_prima", None) else 0),
    ("Origem Materia Prima",    lambda e: e.materia_prima.get_origem_display() if getattr(e, "materia_prima", None) else ""),
]

# index kolom "Total Empregador" (0-based dalam HEADERS, 1-based di Excel)
_EMP_COL_IDX = next(i for i, (h, _) in enumerate(HEADERS) if h == "Total Empregador")


# ===========================================================================
# EXPORT EXCEL
# ===========================================================================

@login_required
@allowed_users(allowed_roles=["admin", "mpms", "XFD"])
def export_excel_mpms(request):
    """
    Download .xlsx data MPMS.

    PERBAIKAN bug MergedCell read-only:
        - Baris 1 (judul) dan baris 2 (filter) TIDAK memakai merge_cells.
          Cukup tulis ke sel A1 / A2 dan set alignment center.
          merge_cells menyebabkan sel B1..Z1 menjadi MergedCell (read-only),
          sehingga penulisan ke sel mana pun di baris itu di luar A1 gagal.
        - Baris total ditulis PER SEL — tidak ada merge yang menabrak
          kolom yang ingin ditulis nilai/formula.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    params   = request.GET
    qs       = _get_filtered_qs(params)
    qs_list  = list(qs)          # evaluasi sekali, hindari double query
    label    = _filter_label(params)
    today    = date.today().strftime("%d-%m-%Y")
    filename = f"MPMS_Export_{today}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Dados MPMS"

    # ── Style ──
    BLUE_DARK = "1F3864"
    WHITE     = "FFFFFF"
    GRAY      = "F2F2F2"

    header_font  = Font(name="Arial", bold=True, color=WHITE, size=10)
    title_font   = Font(name="Arial", bold=True, size=12, color=BLUE_DARK)
    sub_font     = Font(name="Arial", size=9, italic=True, color="555555")
    normal_font  = Font(name="Arial", size=9)
    total_font   = Font(name="Arial", bold=True, size=9, color=BLUE_DARK)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style="thin"),  right=Side(style="thin"),
        top=Side(style="thin"),   bottom=Side(style="thin"),
    )
    header_fill = PatternFill("solid", fgColor=BLUE_DARK)
    total_fill  = PatternFill("solid", fgColor="D6E4F0")

    # ── Baris 1: Judul (TANPA merge_cells) ──
    ws["A1"] = "REKAPITULASAUN DADOS MPMS – MINISTERIU KOMERSIU, INDUSTRIA NO ANBIENTE"
    ws["A1"].font      = title_font
    ws["A1"].alignment = center_align
    ws.row_dimensions[1].height = 22

    # ── Baris 2: Filter & tanggal (TANPA merge_cells) ──
    ws["A2"] = f"Filter: {label}     |     Rai tiha: {today}"
    ws["A2"].font      = sub_font
    ws["A2"].alignment = center_align
    ws.row_dimensions[2].height = 16

    # ── Baris 3: kosong ──
    ws.row_dimensions[3].height = 6

    # ── Baris 4: Header kolom ──
    for col_idx, (col_name, _) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border
    ws.row_dimensions[4].height = 30

    # ── Baris 5 dst: Data ──
    DATA_START = 5
    for row_idx, empresa in enumerate(qs_list, start=1):
        excel_row = row_idx + DATA_START - 1
        row_fill  = PatternFill("solid", fgColor=GRAY) if row_idx % 2 == 0 else PatternFill("solid", fgColor=WHITE)

        for col_idx, (_, accessor) in enumerate(HEADERS, start=1):
            value = row_idx if accessor is None else _safe(accessor, empresa)
            cell  = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.font      = normal_font
            cell.fill      = row_fill
            cell.border    = thin_border
            cell.alignment = center_align if col_idx == 1 else left_align

        ws.row_dimensions[excel_row].height = 16

    # ── Baris total: tulis per-sel, TIDAK merge ──
    total_row      = DATA_START + len(qs_list)   # baris tepat di bawah data terakhir
    data_end_row   = total_row - 1               # baris data terakhir
    emp_excel_col  = _EMP_COL_IDX + 1            # 1-based

    # Sel A total_row: teks "Total Rekord: N"
    c = ws.cell(row=total_row, column=1,
                value=f"Total Rekord: {len(qs_list)}")
    c.font = total_font; c.fill = total_fill; c.alignment = left_align

    # Sel kolom Total Empregador: formula SUM
    emp_letter = get_column_letter(emp_excel_col)
    c = ws.cell(row=total_row, column=emp_excel_col,
                value=f"=SUM({emp_letter}{DATA_START}:{emp_letter}{data_end_row})")
    c.font = total_font; c.fill = total_fill; c.alignment = center_align

    # Isi sel lain di baris total dengan fill (supaya rapi)
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=total_row, column=col_idx)
        if cell.value is None:
            cell.value = ""
        cell.fill   = total_fill
        cell.border = thin_border

    ws.row_dimensions[total_row].height = 18

    # ── Lebar kolom ──
    col_widths = {1:5, 2:20, 3:25, 4:25, 5:8, 6:18,
                  7:15, 8:18, 9:15, 10:15, 11:10}
    for col_idx in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_idx, 14)

    # ── Freeze header ──
    ws.freeze_panes = "A5"

    # ── Auto-filter pada header ──
    ws.auto_filter.ref = f"A4:{get_column_letter(len(HEADERS))}4"

    # ── Stream ──
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _safe(accessor, obj):
    """Panggil accessor, kembalikan string kosong jika gagal."""
    try:
        return accessor(obj)
    except Exception:
        return ""


# ===========================================================================
# EXPORT PDF  (tidak berubah — tidak ada bug di sini)
# ===========================================================================

@login_required
@allowed_users(allowed_roles=["admin", "mpms", "XFD"])
def export_pdf_mpms(request):
    from reportlab.lib           import colors
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib.styles    import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units     import cm
    from reportlab.platypus      import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, HRFlowable,
    )

    params   = request.GET
    qs       = _get_filtered_qs(params)
    qs_list  = list(qs)
    label    = _filter_label(params)
    today    = date.today().strftime("%d/%m/%Y")
    filename = f"MPMS_Export_{date.today().strftime('%d-%m-%Y')}.pdf"

    PDF_COLS = [
        ("No",              None),
        ("Tipo Atividade",  lambda e: str(e.tipo_atividade) if e.tipo_atividade else ""),
        ("Naran Diretor",   lambda e: e.benefisiariu.name if e.benefisiariu else ""),
        ("Naran Kompania",  lambda e: e.company_name or ""),
        ("Sexo",            lambda e: e.benefisiariu.sex if e.benefisiariu else ""),
        ("Munisipiu",       lambda e: str(e.lokalizasaun.municipality) if getattr(e, "lokalizasaun", None) else ""),
        ("Postu Adm.",      lambda e: str(e.lokalizasaun.administrativepost) if getattr(e, "lokalizasaun", None) else ""),
        ("Tinan Hari",      lambda e: e.tinan_hari or ""),
        ("Lisensamentu",    lambda e: e.lisensamentu.get_lisensamentu_display() if getattr(e, "lisensamentu", None) else ""),
        ("Tipu Fundus",     lambda e: e.kapital.get_tipu_fundus_display() if getattr(e, "kapital", None) else ""),
        ("Lukru Brutu/Ano", lambda e: f"${e.kapital.lukru_brutu_ano or 0:,.2f}" if getattr(e, "kapital", None) else ""),
        ("Total Empreg.",   lambda e: e.empregador.total_empregador if getattr(e, "empregador", None) else 0),
        ("Origem Materia",  lambda e: e.materia_prima.get_origem_display() if getattr(e, "materia_prima", None) else ""),
    ]

    buffer = io.BytesIO()
    page_w, page_h = landscape(A3)
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A3),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm,  bottomMargin=1.5*cm,
        title="Dados MPMS",
    )

    styles     = getSampleStyleSheet()
    BLUE       = colors.HexColor("#1F3864")
    GRAY       = colors.HexColor("#F2F2F2")
    cell_style = ParagraphStyle("Cell", parent=styles["Normal"], fontSize=7, leading=9)

    story = [
        Paragraph(
            "REKAPITULASAUN DADOS MPMS<br/>MINISTERIU KOMERSIU, INDUSTRIA NO ANBIENTE",
            ParagraphStyle("T", parent=styles["Heading1"], fontSize=13,
                           textColor=BLUE, spaceAfter=4, alignment=1),
        ),
        Paragraph(
            f"Filter: {label}  |  Rai tiha: {today}",
            ParagraphStyle("S", parent=styles["Normal"], fontSize=8,
                           textColor=colors.HexColor("#555555"), spaceAfter=8, alignment=1),
        ),
        HRFlowable(width="100%", thickness=1, color=BLUE, spaceAfter=8),
    ]

    # Header + data rows
    tbl_data = [[c[0] for c in PDF_COLS]]
    for i, empresa in enumerate(qs_list, start=1):
        row = []
        for col_name, accessor in PDF_COLS:
            if accessor is None:
                row.append(str(i))
            else:
                val = _safe(accessor, empresa)
                row.append(
                    Paragraph(str(val), cell_style) if len(str(val)) > 15 else str(val)
                )
        tbl_data.append(row)

    usable_w   = page_w - 3 * cm
    col_ratios = [0.04, 0.10, 0.12, 0.12, 0.05, 0.09,
                  0.09, 0.06, 0.08, 0.08, 0.08, 0.05, 0.09]
    col_widths = [usable_w * r for r in col_ratios]

    row_bg = [
        ("BACKGROUND", (0, idx), (-1, idx),
         GRAY if idx % 2 == 0 else colors.white)
        for idx in range(1, len(tbl_data))
    ]

    tbl = Table(tbl_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), BLUE),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0), 7),
        ("ALIGN",         (0, 0), (-1, 0), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",      (0, 1), (-1, -1), 7),
        ("ALIGN",         (0, 1), (0, -1),  "CENTER"),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("LINEBELOW",     (0, 0), (-1, 0),  1, BLUE),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING",   (0, 0), (-1, -1), 4),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
        *row_bg,
    ]))

    total_empreg = sum(
        e.empregador.total_empregador if getattr(e, "empregador", None) else 0
        for e in qs_list
    )
    story += [
        tbl,
        Spacer(1, 0.4 * cm),
        Paragraph(
            f"<b>Total Rekord: {len(qs_list)}</b> &nbsp;|&nbsp; "
            f"<b>Total Empregador: {total_empreg}</b> &nbsp;|&nbsp; "
            f"Imprimi tiha: {today}",
            ParagraphStyle("Footer", parent=styles["Normal"],
                           fontSize=8, textColor=BLUE),
        ),
    ]

    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ===========================================================================
# HALAMAN FILTER
# ===========================================================================

@login_required
@allowed_users(allowed_roles=["admin", "mpms", "XFD"])
def export_page_mpms(request):
    params = request.GET
    qs     = _get_filtered_qs(params)
    label  = _filter_label(params)

    context = {
        "title":   "Export Dados MPMS",
        "legend":  "Export Dados MPMS (Excel / PDF)",
        "group":   request.user.groups.all()[0].name,
        "total":   qs.count(),
        "label":   label,
        "params":  params.urlencode(),
        "mun_list":  Municipality.active_objects.order_by("name"),
        "tinan_list": (
            mpmsEmpresa.active_objects
            .values_list("tinan_hari", flat=True)
            .distinct().exclude(tinan_hari=None)
            .order_by("-tinan_hari")
        ),
        "tipo_list": TIPO_ATIVIDADE.objects.order_by("name"),
        "sel_mun":          params.get("mun", ""),
        "sel_tinan":        params.get("tinan", ""),
        "sel_tipo":         params.get("tipo", ""),
        "sel_sexo":         params.get("sexo", ""),
        "sel_lisensamentu": params.get("lisensamentu", ""),
        "sel_tipu_fundus":  params.get("tipu_fundus", ""),
    }
    return render(request, "mpms/export.html", context)