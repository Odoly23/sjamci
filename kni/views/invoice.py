import io
from datetime import date

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render

from config.decorators import allowed_users
from custom.models import Municipality, Year, Faze, Sector
from benefisiariu.models import Benefisiariu
from kni.models import Business, LocBussiness, Program, Employee, Finance, \
                       BusinessBaseline, BusinessMonitoring


def _get_filtered_qs(params):
    qs = (
        Benefisiariu.active_objects
        .filter(Pnegosiu__program_type__name="KNI")
        .select_related("status")
        .prefetch_related(
            "locnegosiu__municipality",
            "locnegosiu__administrativepost",
            "locnegosiu__village",
            "Pnegosiu__program_type",
            "Pnegosiu__faze",
            "Pnegosiu__year",
            "Pnegosiu__status",
            "Pnegosiu__t_apoiu",
            "Pnegosiu__t_fundus",
            "negosiu__sector",
            "negosiu__category",
            "negosiu__size",
            "negosiu__employee_set",
            "negosiu__finance_set",
            "negosiu__baseline",
            "negosiu__monitorings",
        )
        .order_by("locnegosiu__municipality__name", "name")
        .distinct()
    )
    mun    = params.get("mun",    "").strip()
    year   = params.get("year",   "").strip()
    faze   = params.get("faze",   "").strip()
    sector = params.get("sector", "").strip()
    sexo   = params.get("sexo",   "").strip()
    status = params.get("status", "").strip()

    if mun:    qs = qs.filter(locnegosiu__municipality__name__icontains=mun)
    if year:   qs = qs.filter(Pnegosiu__year__year=year)
    if faze:   qs = qs.filter(Pnegosiu__faze__name__icontains=faze)
    if sector: qs = qs.filter(negosiu__sector__name__icontains=sector)
    if sexo:   qs = qs.filter(sex__iexact=sexo)
    if status: qs = qs.filter(Pnegosiu__status__name__iexact=status)

    return qs


def _filter_label(params):
    parts = []
    if params.get("mun"):    parts.append(f"Munisipiu: {params['mun']}")
    if params.get("year"):   parts.append(f"Tinan: {params['year']}")
    if params.get("faze"):   parts.append(f"Faze: {params['faze']}")
    if params.get("sector"): parts.append(f"Setor: {params['sector']}")
    if params.get("sexo"):   parts.append(f"Genero: {params['sexo']}")
    if params.get("status"): parts.append(f"Status: {params['status']}")
    return " | ".join(parts) if parts else "Dadus Tomak KNI (Hotu)"


def _biz(benef):
    return benef.negosiu.first()

def _loc(benef):
    return benef.locnegosiu.first()

def _prog(benef):
    return benef.Pnegosiu.filter(program_type__name="KNI").first()

def _emp(benef):
    biz = _biz(benef)
    return biz.employee_set.first() if biz else None

def _fin(benef):
    biz = _biz(benef)
    return biz.finance_set.first() if biz else None

def _baseline(benef):
    biz = _biz(benef)
    if not biz:
        return None
    try:
        return biz.baseline
    except Exception:
        return None

def _mon(benef):
    biz = _biz(benef)
    return biz.monitorings.first() if biz else None

def _safe(fn):
    try:
        v = fn()
        return v if v is not None else ""
    except Exception:
        return ""


KNI_HEADERS = [
    ("No",                    None),
    ("Naran Kompletu",        lambda b: b.name or ""),
    ("Genero",                lambda b: b.sex or ""),
    ("No. Tlp.",              lambda b: b.phone or ""),
    ("Nivel Edukasaun",       lambda b: b.get_nivel_edukasaun_display() if b.nivel_edukasaun else ""),
    ("Email / Website",       lambda b: b.email_website or ""),
    ("Naran Negocio/Empresa", lambda b: _safe(lambda: _biz(b).name or "")),
    ("Ideia Negocio",         lambda b: _safe(lambda: _biz(b).idea or "")),
    ("Seitor Prinsipal",      lambda b: _safe(lambda: str(_biz(b).sector) if _biz(b) and _biz(b).sector else "")),
    ("Kategoria",             lambda b: _safe(lambda: str(_biz(b).category) if _biz(b) and _biz(b).category else "")),
    ("Tamañu Negosiu",        lambda b: _safe(lambda: str(_biz(b).size) if _biz(b) and _biz(b).size else "")),
    ("Total Trabalhador",     lambda b: _safe(lambda: _emp(b).total if _emp(b) else 0)),
    ("Feto",                  lambda b: _safe(lambda: _emp(b).female if _emp(b) else 0)),
    ("Mane",                  lambda b: _safe(lambda: _emp(b).male if _emp(b) else 0)),
    ("Munisipiu",             lambda b: _safe(lambda: str(_loc(b).municipality) if _loc(b) and _loc(b).municipality else "")),
    ("Postu-Administrativu",  lambda b: _safe(lambda: str(_loc(b).administrativepost) if _loc(b) and _loc(b).administrativepost else "")),
    ("Suku",                  lambda b: _safe(lambda: str(_loc(b).village) if _loc(b) and _loc(b).village else "")),
    ("Aldeia",                lambda b: _safe(lambda: _loc(b).aldeia or "" if _loc(b) else "")),
    ("Faze",                  lambda b: _safe(lambda: str(_prog(b).faze) if _prog(b) and _prog(b).faze else "")),
    ("Tinan",                 lambda b: _safe(lambda: str(_prog(b).year) if _prog(b) and _prog(b).year else "")),
    ("KNI (Faze+Tinan)",      lambda b: _safe(lambda: f"{_prog(b).faze} {_prog(b).year}" if _prog(b) and _prog(b).faze and _prog(b).year else "")),
    ("Tipu Apoiu",            lambda b: _safe(lambda: str(_prog(b).t_apoiu) if _prog(b) and _prog(b).t_apoiu else "")),
    ("Tipu Fundus Kapital",   lambda b: _safe(lambda: str(_prog(b).t_fundus) if _prog(b) and _prog(b).t_fundus else "")),
    ("Status Programa",       lambda b: _safe(lambda: str(_prog(b).status) if _prog(b) and _prog(b).status else "")),
    ("Montante Aprova ($)",   lambda b: _safe(lambda: float(_prog(b).approved_amount or 0) if _prog(b) else 0)),
    ("Montante Apoiu ($)",    lambda b: _safe(lambda: float(_prog(b).amount or 0) if _prog(b) else 0)),
    ("Total Orsamento ($)",   lambda b: _safe(lambda: float(_fin(b).budget or 0) if _fin(b) else 0)),
]

BASELINE_HEADERS = [
    ("No",                          None),
    ("Naran Kompletu",              lambda b: b.name or ""),
    ("Naran Negocio",               lambda b: _safe(lambda: _biz(b).name or "")),
    ("Rendimentu Loron Antes ($)",  lambda b: _safe(lambda: float(_baseline(b).daily_income_before or 0) if _baseline(b) else 0)),
    ("Rendimentu Fulan Antes ($)",  lambda b: _safe(lambda: float(_baseline(b).monthly_income_before or 0) if _baseline(b) else 0)),
    ("Rendimentu Tinan Antes ($)",  lambda b: _safe(lambda: float(_baseline(b).yearly_income_before or 0) if _baseline(b) else 0)),
    ("Trabalhador Antes",           lambda b: _safe(lambda: _baseline(b).employee_before or 0 if _baseline(b) else 0)),
    ("Assets Antes ($)",            lambda b: _safe(lambda: float(_baseline(b).asset_before or 0) if _baseline(b) else 0)),
    ("Vendas Antes ($)",            lambda b: _safe(lambda: float(_baseline(b).sales_before or 0) if _baseline(b) else 0)),
    ("Observasaun",                 lambda b: _safe(lambda: _baseline(b).note or "" if _baseline(b) else "")),
]

MON_HEADERS = [
    ("No",                    None),
    ("Naran Kompletu",        lambda b: b.name or ""),
    ("Naran Negocio",         lambda b: _safe(lambda: _biz(b).name or "")),
    ("Tinan",                 lambda b: _safe(lambda: str(_mon(b).year) if _mon(b) and _mon(b).year else "")),
    ("Fulan",                 lambda b: _safe(lambda: _mon(b).month or "" if _mon(b) else "")),
    ("Data Monitorizasaun",   lambda b: _safe(lambda: _mon(b).monitoring_date.strftime("%d/%m/%Y") if _mon(b) else "")),
    ("Rendimentu Loron ($)",  lambda b: _safe(lambda: float(_mon(b).daily_income or 0) if _mon(b) else 0)),
    ("Rendimentu Fulan ($)",  lambda b: _safe(lambda: float(_mon(b).monthly_income or 0) if _mon(b) else 0)),
    ("Rendimentu Tinan ($)",  lambda b: _safe(lambda: float(_mon(b).yearly_income or 0) if _mon(b) else 0)),
    ("Total Vendas ($)",      lambda b: _safe(lambda: float(_mon(b).total_sales or 0) if _mon(b) else 0)),
    ("Total Assets ($)",      lambda b: _safe(lambda: float(_mon(b).total_assets or 0) if _mon(b) else 0)),
    ("Total Trabalhador",     lambda b: _safe(lambda: _mon(b).total_employee or 0 if _mon(b) else 0)),
    ("Cresimentu (%)",        lambda b: _safe(lambda: _mon(b).growth_percentage or 0 if _mon(b) else 0)),
    ("Status Monitorizasaun", lambda b: _safe(lambda: _mon(b).monitoring_status or "" if _mon(b) else "")),
    ("Status Verifikasaun",   lambda b: _safe(lambda: _mon(b).verification_status or "" if _mon(b) else "")),
    ("Fonte Dadus",           lambda b: _safe(lambda: _mon(b).source_data or "" if _mon(b) else "")),
    ("Observasaun",           lambda b: _safe(lambda: _mon(b).note or "" if _mon(b) else "")),
]


@login_required
@allowed_users(allowed_roles=["admin", "KNI", "XFD"])
def export_excel_kni(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    params   = request.GET
    qs       = _get_filtered_qs(params)
    benef_list = list(qs)
    label    = _filter_label(params)
    today    = date.today().strftime("%d-%m-%Y")
    filename = f"KNI_Export_{today}.xlsx"
    wb = Workbook()
    BLUE   = "1F3864"
    GREEN  = "1A5276"
    PURPLE = "4A235A"
    WHITE  = "FFFFFF"
    GRAY   = "F2F2F2"
    LBLUE  = "D6E4F0"
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin   = Border(
        left=Side(style="thin"),  right=Side(style="thin"),
        top=Side(style="thin"),   bottom=Side(style="thin"),
    )
    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)
    def _write_sheet(ws, headers, data, title_text, hdr_hex):
        ws["A1"] = title_text
        ws["A1"].font      = Font(name="Arial", bold=True, size=12, color=hdr_hex)
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 22
        ws["A2"] = f"Filter: {label}  |  Rai tiha: {today}"
        ws["A2"].font      = Font(name="Arial", size=9, italic=True, color="555555")
        ws["A2"].alignment = center
        ws.row_dimensions[2].height = 16
        ws.row_dimensions[3].height = 6
        hdr_fill = fill(hdr_hex)
        hdr_font = Font(name="Arial", bold=True, color=WHITE, size=10)
        for col, (name, _) in enumerate(headers, 1):
            c = ws.cell(row=4, column=col, value=name)
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = center; c.border = thin
        ws.row_dimensions[4].height = 28
        DATA_START = 5
        nfont = Font(name="Arial", size=9)
        for i, benef in enumerate(data, 1):
            row_fill = fill(GRAY) if i % 2 == 0 else fill(WHITE)
            for col, (_, acc) in enumerate(headers, 1):
                val = i if acc is None else _safe(lambda a=acc, b=benef: a(b))
                c = ws.cell(row=DATA_START + i - 1, column=col, value=val)
                c.font = nfont; c.fill = row_fill
                c.border = thin
                c.alignment = center if col == 1 else left
            ws.row_dimensions[DATA_START + i - 1].height = 16
        total_row = DATA_START + len(data)
        tfont = Font(name="Arial", bold=True, size=9, color=hdr_hex)
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=total_row, column=col, value="")
            c.fill = fill(LBLUE); c.border = thin
        ws.cell(row=total_row, column=1,
                value=f"Total Rekord: {len(data)}").font = tfont
        for col, (name, _) in enumerate(headers, 1):
            ltr = get_column_letter(col)
            ws.column_dimensions[ltr].width = max(10, min(len(name) + 4, 28))

        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}4"

    ws1 = wb.active
    ws1.title = "Dados KNI"
    _write_sheet(
        ws1, KNI_HEADERS, benef_list,
        "REKAPITULASAUN DADOS KNI – KOMPETISAUN NEGOSIU INOVATIVU",
        BLUE,
    )

    ws2 = wb.create_sheet("Baseline")
    with_baseline = [b for b in benef_list if _baseline(b)]
    _write_sheet(
        ws2, BASELINE_HEADERS, with_baseline,
        "DADUS INISIÁL NEGÓSIU (BASELINE) – KNI",
        GREEN,
    )

    ws3 = wb.create_sheet("Monitorizasaun")
    with_mon = [b for b in benef_list if _mon(b)]
    _write_sheet(
        ws3, MON_HEADERS, with_mon,
        "DADUS MONITORIZASAUN NEGÓSIU – KNI",
        PURPLE,
    )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
@allowed_users(allowed_roles=["admin", "KNI", "XFD"])
def export_pdf_kni(request):
    from reportlab.lib           import colors
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib.styles    import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units     import cm
    from reportlab.platypus      import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, HRFlowable,
    )

    params     = request.GET
    qs         = _get_filtered_qs(params)
    benef_list = list(qs)
    label      = _filter_label(params)
    today      = date.today().strftime("%d/%m/%Y")
    filename   = f"KNI_Export_{date.today().strftime('%d-%m-%Y')}.pdf"

    PDF_COLS = [
        ("No",             None),
        ("Naran Kompletu", lambda b: b.name or ""),
        ("Genero",         lambda b: b.sex or ""),
        ("Naran Negocio",  lambda b: _safe(lambda: _biz(b).name or "")),
        ("Ideia Negocio",  lambda b: _safe(lambda: _biz(b).idea or "")),
        ("Seitor",         lambda b: _safe(lambda: str(_biz(b).sector) if _biz(b) and _biz(b).sector else "")),
        ("Mane",           lambda b: _safe(lambda: _emp(b).male if _emp(b) else 0)),
        ("Feto",           lambda b: _safe(lambda: _emp(b).female if _emp(b) else 0)),
        ("Total Trab.",    lambda b: _safe(lambda: _emp(b).total if _emp(b) else 0)),
        ("Munisipiu",      lambda b: _safe(lambda: str(_loc(b).municipality) if _loc(b) and _loc(b).municipality else "")),
        ("Postu",          lambda b: _safe(lambda: str(_loc(b).administrativepost) if _loc(b) and _loc(b).administrativepost else "")),
        ("Faze",           lambda b: _safe(lambda: str(_prog(b).faze) if _prog(b) and _prog(b).faze else "")),
        ("Tinan",          lambda b: _safe(lambda: str(_prog(b).year) if _prog(b) and _prog(b).year else "")),
        ("Status",         lambda b: _safe(lambda: str(_prog(b).status) if _prog(b) and _prog(b).status else "")),
        ("Orsamento ($)",  lambda b: _safe(lambda: float(_fin(b).budget or 0) if _fin(b) else 0)),
    ]

    buffer = io.BytesIO()
    page_w, _ = landscape(A3)
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A3),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm,  bottomMargin=1.5*cm,
        title="Dados KNI",
    )

    styles     = getSampleStyleSheet()
    BLUE       = colors.HexColor("#1F3864")
    GRAY       = colors.HexColor("#F2F2F2")
    cell_style = ParagraphStyle("Cell", parent=styles["Normal"], fontSize=7, leading=9)

    # Hitung total footer
    total_emp = sum(_safe(lambda b=b: _emp(b).total if _emp(b) else 0) or 0 for b in benef_list)
    total_ors = sum(_safe(lambda b=b: float(_fin(b).budget or 0) if _fin(b) else 0) or 0 for b in benef_list)

    story = [
        Paragraph(
            "REKAPITULASAUN DADOS KNI<br/>KOMPETISAUN NEGOSIU INOVATIVU",
            ParagraphStyle("T", parent=styles["Heading1"], fontSize=13,
                           textColor=BLUE, spaceAfter=4, alignment=1),
        ),
        Paragraph(
            f"Filter: {label}  |  Rai tiha: {today}",
            ParagraphStyle("S", parent=styles["Normal"], fontSize=8,
                           textColor=colors.HexColor("#555555"),
                           spaceAfter=8, alignment=1),
        ),
        HRFlowable(width="100%", thickness=1, color=BLUE, spaceAfter=8),
    ]

    # Build table
    tbl_data = [[c[0] for c in PDF_COLS]]
    for i, benef in enumerate(benef_list, 1):
        row = []
        for col_name, acc in PDF_COLS:
            if acc is None:
                row.append(str(i))
            else:
                val = str(_safe(lambda a=acc, b=benef: a(b)))
                row.append(Paragraph(val, cell_style) if len(val) > 18 else val)
        tbl_data.append(row)

    usable_w   = page_w - 3 * cm
    col_ratios = [0.04, 0.12, 0.05, 0.12, 0.12, 0.07,
                  0.04, 0.04, 0.05, 0.08, 0.08, 0.05, 0.05, 0.07, 0.07]
    col_widths = [usable_w * r for r in col_ratios]

    row_bg = [
        ("BACKGROUND", (0, idx), (-1, idx),
         GRAY if idx % 2 == 0 else colors.white)
        for idx in range(1, len(tbl_data))
    ]

    tbl = Table(tbl_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), BLUE),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0), 7),
        ("ALIGN",         (0, 0), (-1, 0), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",      (0, 1), (-1, -1), 7),
        ("ALIGN",         (0, 1), (0, -1),  "CENTER"),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("LINEBELOW",     (0, 0), (-1, 0),  1,   BLUE),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING",   (0, 0), (-1, -1), 4),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
        *row_bg,
    ]))

    story += [
        tbl,
        Spacer(1, 0.4 * cm),
        Paragraph(
            f"<b>Total Rekord: {len(benef_list)}</b> &nbsp;|&nbsp; "
            f"<b>Total Trabalhador: {total_emp}</b> &nbsp;|&nbsp; "
            f"<b>Total Orsamento: ${total_ors:,.2f}</b> &nbsp;|&nbsp; "
            f"Imprimi tiha: {today}",
            ParagraphStyle("Footer", parent=styles["Normal"],
                           fontSize=8, textColor=BLUE),
        ),
    ]

    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
@allowed_users(allowed_roles=["admin", "KNI", "XFD"])
def export_page_kni(request):
    params = request.GET
    qs     = _get_filtered_qs(params)
    label  = _filter_label(params)

    context = {
        "title":   "Export Dados KNI",
        "legend":  "Export Dados KNI (Excel / PDF)",
        "group":   request.user.groups.all()[0].name,
        "total":   qs.count(),
        "label":   label,
        "params":  params.urlencode(),
        "mun_list":    Municipality.active_objects.order_by("name"),
        "year_list":   Year.active_objects.order_by("-year"),
        "faze_list":   Faze.active_objects.exclude(
                           name__in=["KREDITU", "mpms", "manufatur"]
                       ).order_by("name"),
        "sector_list": Sector.active_objects.order_by("name"),
        "sel_mun":    params.get("mun",    ""),
        "sel_year":   params.get("year",   ""),
        "sel_faze":   params.get("faze",   ""),
        "sel_sector": params.get("sector", ""),
        "sel_sexo":   params.get("sexo",   ""),
        "sel_status": params.get("status", ""),
        "excel_cols": [
            "Naran Kompletu", "Genero", "No. Tlp.", "Nivel Edukasaun",
            "Email / Website", "Naran Negocio/Empresa", "Ideia Negocio",
            "Seitor Prinsipal", "Kategoria", "Tamañu Negosiu",
            "Total Trabalhador", "Feto", "Mane",
            "Munisipiu", "Postu-Administrativu", "Suku", "Aldeia",
            "Faze", "Tinan", "KNI (Faze+Tinan)", "Tipu Apoiu",
            "Tipu Fundus Kapital", "Status Programa",
            "Montante Aprova ($)", "Montante Apoiu ($)", "Total Orsamento ($)",
        ],
    }
    return render(request, "Dash/export_kni.html", context)

